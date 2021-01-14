from openpyxl import load_workbook
import os
import requests
from urllib.error import HTTPError
import json
import time

# Variables
BASEURL = 'https://ttotenberg.das.perforce.com:8443/helix-alm/api/v0/'
APIKEY = 'df4b2b246a00d5e6459620d99dd970a4aa38f4a46542c3c5c5148185e0b4dd5d:e7758a2065a0ee841da370221be9b9eb26c9a8afb2d74e78d45d8be6cc42a6ea'
PROJECTID = 28
WORKBOOK = 'Sample Data.xlsx'
HELIXFIELD = 'LinkReference'

# Define some useful variables for ease of typing later. Also convert testcase and testrun to testCase and testRun for json purposes.
url = BASEURL + str(PROJECTID) + '/'
wb = load_workbook(WORKBOOK)
ws = wb.active

for col in ws:
  for cell in col:
    if cell.value == 'testcase':
      cell.value = 'testCase'
    elif cell.value == 'testrun':
      cell.value = 'testRun'
wb.save(WORKBOOK)

# Get your Helix ALM API Bearer token, and set values for authentication used later
blankPayload  = {}
headers = {'Authorization': 'APIKey ' + APIKEY}
token = ''

def getToken():
  try:
    response = requests.request("GET", url + 'token', headers=headers, data = blankPayload, verify=False)
    response.raise_for_status()
    jsonResponse = response.json()
    global token
    token = jsonResponse["accessToken"]

  except HTTPError as http_err:
    print(f'HTTP error occurred: {http_err}')
  except Exception as err:
    print(f'Other error occurred: {err}')

getToken()
headers = {'Authorization': 'Bearer ' + token}

# Define your Excel columns
ltc = 1 # ltc = your link type column
lnc = 2 # lnc = link name column
ppitc = 3 # ppitc = parent or peer item type column
ppidc = 4 # ppidc = parent or peer identifier column
cpitc = 5 # cpitc = child or peer item type column
cpidc = 6 # cpidc = child or peer identifier column

# Define your starting Excel row
cr = 2 # cr = current row  

while ws.cell(row=cr, column=ltc).value != None:
  if ws.cell(row=cr, column=ltc).value != None:
    # Define the cell values
    # Link Type
    linkType = ws.cell(row=cr, column=ltc).value
    # Link Name
    linkName = ws.cell(row=cr, column=lnc).value
    # Parent/Peer Item Type
    parentPeerType = ws.cell(row=cr, column=ppitc).value
    # Parent/Peer Identifier
    parentPeerId = ws.cell(row=cr, column=ppidc).value
    # Child/Peer Item Type
    childPeerType = ppit = ws.cell(row=cr, column=cpitc).value
    # Child/Peer Identifier
    childPeerId = ppit = ws.cell(row=cr, column=cpidc).value

  # Create a Parent/Child Link
  while ws.cell(row=cr, column=ltc).value != None:

    # Check rate limits, and sleep if the rate is below 10
    response = requests.request("GET", url + 'ratelimits', headers=headers, data = blankPayload, verify=False)
    responseRateLimitRemaining = response.headers['X-RateLimit-Remaining']
    print('The remaining rate limit is: ' + str(responseRateLimitRemaining))

    if int(responseRateLimitRemaining) <= 5:
      time.sleep(60) 
    else:
      pass

    if linkType == 'ParentChild':
      
      # Get the Parent/Peer ID based on the 'HELIXFIELD' identifier
      if parentPeerType == 'testRun':
        response = requests.request("GET", url + parentPeerType +'s' + '?search=' + HELIXFIELD + '="' + str(parentPeerId) + '"' + '&fields=Test Run Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        parentPeerNumber = jsonResponse[parentPeerType + 's'][0]["fields"][0]["integer"]
      elif parentPeerType != 'testRun':
        response = requests.request("GET", url + parentPeerType +'s' + '?search=' + HELIXFIELD + '="' + str(parentPeerId) + '"' + '&fields=Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        parentPeerNumber = jsonResponse[parentPeerType + 's'][0]["fields"][0]["integer"]
      else:
        pass
      
      # Get the Child/Peer ID based on the 'HELIXFIELD' identifier
      if childPeerType == 'testRun':
        response = requests.request("GET", url + childPeerType +'s' + '?search=' + HELIXFIELD + '="' + str(childPeerId) + '"' '&fields=Test Run Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        childPeerNumber = jsonResponse[childPeerType + 's'][0]["fields"][0]["integer"]
      elif childPeerType != 'testRun':
        response = requests.request("GET", url + childPeerType +'s' + '?search=' + HELIXFIELD + '="' + str(childPeerId) + '"' '&fields=Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        childPeerNumber = jsonResponse[childPeerType + 's'][0]["fields"][0]["integer"]
      else:
        pass
      
      print('The Parent ID is: ' + str(parentPeerNumber) + ' and the Child ID is: ' + str(childPeerNumber))

      # Link the Parent and Child Together
      dictionaryPayload = {"linksData":[{"linkDefinition":{"name": linkName},"parentChildren": {"parent": {"itemID": parentPeerNumber,"itemType": parentPeerType + 's'},"children": [{"itemID": childPeerNumber,"itemType": childPeerType + 's'}]}}]}
      payload = json.dumps(dictionaryPayload)
      print(payload)
      jsonHeaders = {'Authorization': 'Bearer ' + token,'Content-Type': 'application/json'}
      
      
      try:
        response = requests.post(url + parentPeerType + 's/' + str(parentPeerNumber) + '/links', headers=jsonHeaders, data = payload, verify=False)
        response.raise_for_status()

      except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
      except Exception as err:
        print(f'Other error occurred: {err}')
    else:
      pass

    # Create a Peer to Peer link
    if linkType == 'Peers':

      # Get the Parent/Peer ID based on the 'HELIXFIELD' identifier
      if parentPeerType == 'testRun':
        response = requests.request("GET", url + parentPeerType +'s' + '?search=' + HELIXFIELD + '=' + str(parentPeerId) + '&fields=Test Run Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        parentPeerNumber = jsonResponse[parentPeerType + 's'][0]["fields"][0]["integer"]
      elif parentPeerType != 'testRun':
        response = requests.request("GET", url + parentPeerType +'s' + '?search=' + HELIXFIELD + '=' + str(parentPeerId) + '&fields=Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        #print('The ParentPeerNumber response is: ' + str(jsonResponse))
        parentPeerNumber = jsonResponse[parentPeerType + 's'][0]["fields"][0]["integer"]
      else:
        pass
      
      # Get the Child/Peer ID based on the 'HELIXFIELD' identifier
      if childPeerType == 'testRun':
        response = requests.request("GET", url + childPeerType +'s' + '?search=' + HELIXFIELD + '=' + str(childPeerId) + '&fields=Test Run Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        childPeerNumber = jsonResponse[childPeerType + 's'][0]["fields"][0]["integer"]
      elif childPeerType != 'testRun':
        response = requests.request("GET", url + childPeerType +'s' + '?search=' + HELIXFIELD + '=' + str(childPeerId) + '&fields=Number', headers=headers, data = blankPayload, verify=False)
        jsonResponse = response.json()
        childPeerNumber = jsonResponse[childPeerType + 's'][0]["fields"][0]["integer"]
      else:
        pass

      print('The two linked items are: ' + str(parentPeerNumber) + ' and ' + str(childPeerNumber))
      
      # Link the peers together
      dictionaryPayload = { "linksData": [ { "linkDefinition": { "name": linkName }, "type": "peers", "peers": [ { "itemID": parentPeerNumber, "itemType": parentPeerType + 's' }, { "itemID": childPeerNumber, "itemType": childPeerType + 's' } ] } ] }
      payload = json.dumps(dictionaryPayload)
      jsonHeaders = {'Authorization': 'Bearer ' + token,'Content-Type': 'application/json'}

      try:
        response = requests.post(url + parentPeerType + 's/' + str(parentPeerNumber) + '/links', headers=jsonHeaders, data = payload, verify=False)
        response.raise_for_status()

      except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
      except Exception as err:
        print(f'Other error occurred: {err}')
    else:
      pass
    cr = cr + 1
    if ws.cell(row=cr, column=ltc).value == None:
      break
    break
